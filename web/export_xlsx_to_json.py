# -*- coding: utf-8 -*-
"""從 地點報表.xlsx 匯出 places.json 供網頁使用。"""
import json
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    raise

# 報表在專案根目錄
XLSX_PATH = Path(__file__).resolve().parent.parent / "地點報表.xlsx"
OUT_PATH = Path(__file__).resolve().parent / "data" / "places.json"

HEADER_MAP = {
    "A名稱": "name",
    "B地址": "addressUrl",
    "C分店類型": "type",
    "D相關介紹網頁": "links",
    "特色": "specialty",
    "價錢範圍(港幣)": "priceHkd",
    "開店時間": "hours",
    "分數": "rating",
    "優點": "pros",
    "缺點": "cons",
    "距離": "distance",
    "備註": "remarks",
    "最近交通": "transport",
    "TAG": "tags",
    "相片1": "photo1Url",
    "相片2": "photo2Url",
}


def cell_str(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def cell_hyperlink(cell):
    """若儲存格有超連結則回傳 URL，否則回傳 None"""
    try:
        if cell.hyperlink and getattr(cell.hyperlink, "target", None):
            return cell.hyperlink.target
    except Exception:
        pass
    return None


def main():
    if not XLSX_PATH.exists():
        print(f"找不到 {XLSX_PATH}")
        return
    wb = load_workbook(XLSX_PATH, data_only=True)  # 值用 data_only；hyperlink 仍可讀
    ws = wb.active
    # 第一列為標題，建立 欄位名 -> 欄索引
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = cell_str(ws.cell(row=1, column=c))
        if h in HEADER_MAP:
            col_map[HEADER_MAP[h]] = c
    # 需讀取超連結的欄位（Excel 存的是 URL）
    url_keys = ("addressUrl", "photo1Url", "photo2Url")
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        raw = {}
        for key, col in col_map.items():
            cell = ws.cell(row=row_idx, column=col)
            if key in url_keys:
                link = cell_hyperlink(cell)
                raw[key] = link if link else cell_str(cell)
            else:
                raw[key] = cell_str(cell)
        # 相片1 若為空可用 addressUrl（地圖）
        if not raw.get("photo1Url") and raw.get("addressUrl"):
            raw["photo1Url"] = raw["addressUrl"]
        # TAG 字串轉陣列（頓號、逗號分隔）
        tag_str = raw.get("tags") or ""
        raw["tags"] = [t.strip() for t in re.split(r"[、,，]", tag_str) if t.strip()]
        raw["id"] = row_idx - 2  # 0-based id
        rows.append(raw)
    wb.close()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f"已匯出 {len(rows)} 筆至 {OUT_PATH}")


if __name__ == "__main__":
    main()
