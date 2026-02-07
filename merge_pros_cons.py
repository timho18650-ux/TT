# -*- coding: utf-8 -*-
"""將「優點／缺點」CSV 合併進已產生的 地點報表.xlsx。

CSV 格式：第一欄為店名（與 Excel A名稱 對應），第二欄優點，第三欄缺點。
可含標題列，例如：
  名稱,優點,缺點
  Canary Massage,服務專業環境佳,...,
"""
import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    raise

EXCEL_NAME = "地點報表.xlsx"
# Excel 欄位索引（1-based）：1=A名稱, 9=優點, 10=缺點
COL_NAME, COL_PRO, COL_CON = 1, 9, 10


def main():
    base = Path(__file__).parent
    xlsx_path = base / EXCEL_NAME
    csv_path = base / "優缺點.csv"

    if not xlsx_path.exists():
        print(f"找不到 {EXCEL_NAME}，請先執行 kml_to_excel.py 產生報表。")
        return
    if not csv_path.exists():
        print(f"請建立 {csv_path.name}，格式：名稱,優點,缺點（第一列可為標題）")
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active
    # 讀取 CSV：假設 名稱, 優點, 缺點
    by_name = {}
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 3:
                continue
            name, pro, con = row[0].strip(), row[1].strip(), row[2].strip()
            if name in ("名稱", "A名稱", "name", "店名"):
                continue
            by_name[name] = (pro, con)

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=COL_NAME)
        name = (name_cell.value or "").strip()
        if name in by_name:
            pro, con = by_name[name]
            ws.cell(row=row_idx, column=COL_PRO, value=pro)
            ws.cell(row=row_idx, column=COL_CON, value=con)
            updated += 1

    wb.save(xlsx_path)
    print(f"已更新 {updated} 筆優點／缺點至 {EXCEL_NAME}")


if __name__ == "__main__":
    main()
