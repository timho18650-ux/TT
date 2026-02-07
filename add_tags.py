# -*- coding: utf-8 -*-
"""從 tags.csv 讀取每間地點的 TAG（最多 6 個通用標籤），寫入 地點報表.xlsx 新增的「TAG」欄。"""
import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    raise

EXCEL_NAME = "地點報表.xlsx"
COL_NAME = 1


def main():
    base = Path(__file__).parent
    xlsx_path = base / EXCEL_NAME
    csv_path = base / "tags.csv"
    if not xlsx_path.exists():
        print(f"找不到 {EXCEL_NAME}，請先執行 kml_to_excel.py。")
        return
    if not csv_path.exists():
        print(f"請建立 {csv_path.name}，格式：名稱,TAG（TAG 以頓號分隔，每間最多 6 個）")
        return

    by_name = {}
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if len(row) < 2:
                continue
            name, tag_str = row[0].strip(), row[1].strip()
            if name in ("名稱", "A名稱", "name"):
                continue
            by_name[name] = tag_str

    wb = load_workbook(xlsx_path)
    ws = wb.active
    max_col = ws.max_column
    # 檢查是否已有 TAG 欄
    tag_col = None
    for c in range(1, max_col + 1):
        if (ws.cell(row=1, column=c).value or "").strip() == "TAG":
            tag_col = c
            break
    if tag_col is None:
        tag_col = max_col + 1
        ws.cell(row=1, column=tag_col, value="TAG").font = Font(bold=True)
        ws.cell(row=1, column=tag_col).alignment = Alignment(wrap_text=True, vertical="top")

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        name = (ws.cell(row=row_idx, column=COL_NAME).value or "").strip()
        if name in by_name:
            ws.cell(row=row_idx, column=tag_col, value=by_name[name])
            ws.cell(row=row_idx, column=tag_col).alignment = Alignment(wrap_text=True, vertical="top")
            updated += 1

    ws.column_dimensions[ws.cell(row=1, column=tag_col).column_letter].width = 38
    wb.save(xlsx_path)
    print(f"已為 {updated} 筆地點寫入 TAG 至 {EXCEL_NAME}")


if __name__ == "__main__":
    main()
