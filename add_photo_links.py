# -*- coding: utf-8 -*-
"""為地點報表.xlsx 每筆地點加入「相片1」「相片2」兩欄相片超連結（不下載檔案）。
相片1 = Google 地圖該地點（可看地圖與用戶上傳相片）
相片2 = Google 圖片搜尋該地點
"""
from urllib.parse import quote
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    raise

EXCEL_NAME = "地點報表.xlsx"
COL_NAME = 1
COL_ADDRESS = 2  # B 欄為 Google 地圖連結


def google_image_search_url(name):
    """產生 Google 圖片搜尋「地點名 Bangkok」的網址"""
    q = f"{name} Bangkok"
    return f"https://www.google.com/search?q={quote(q)}&tbm=isch"


def main():
    base = Path(__file__).parent
    xlsx_path = base / EXCEL_NAME
    if not xlsx_path.exists():
        print(f"找不到 {EXCEL_NAME}，請先執行 kml_to_excel.py。")
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active
    max_col = ws.max_column
    header_row = 1

    # 找是否已有 相片1、相片2
    col_photo1 = col_photo2 = None
    for c in range(1, max_col + 1):
        v = (ws.cell(row=header_row, column=c).value or "").strip()
        if v == "相片1":
            col_photo1 = c
        elif v == "相片2":
            col_photo2 = c

    # 若沒有則在最後新增兩欄
    if col_photo1 is None:
        col_photo1 = max_col + 1
        ws.cell(row=header_row, column=col_photo1, value="相片1").font = Font(bold=True)
    if col_photo2 is None:
        col_photo2 = (col_photo1 + 1) if col_photo2 is None else col_photo2
        if col_photo2 == col_photo1:
            col_photo2 = max_col + 2
        if ws.cell(row=header_row, column=col_photo2).value is None or (ws.cell(row=header_row, column=col_photo2).value or "").strip() != "相片2":
            ws.cell(row=header_row, column=col_photo2, value="相片2").font = Font(bold=True)

    for row_idx in range(2, ws.max_row + 1):
        name = (ws.cell(row=row_idx, column=COL_NAME).value or "").strip()
        maps_url = (ws.cell(row=row_idx, column=COL_ADDRESS).value or "").strip()
        if not name:
            continue
        # 相片1：Google 地圖（可看該地點與用戶相片）
        c1 = ws.cell(row=row_idx, column=col_photo1, value="相片1")
        if maps_url and maps_url.startswith("http"):
            c1.hyperlink = maps_url
            c1.font = Font(color="0563C1", underline="single")
        c1.alignment = Alignment(vertical="top")
        # 相片2：Google 圖片搜尋
        img_url = google_image_search_url(name)
        c2 = ws.cell(row=row_idx, column=col_photo2, value="相片2")
        c2.hyperlink = img_url
        c2.font = Font(color="0563C1", underline="single")
        c2.alignment = Alignment(vertical="top")

    for col in (col_photo1, col_photo2):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 10
    wb.save(xlsx_path)
    print(f"已為 {ws.max_row - 1} 筆地點加入相片1／相片2 超連結至 {EXCEL_NAME}")


if __name__ == "__main__":
    main()
