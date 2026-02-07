# -*- coding: utf-8 -*-
"""從 地圖.kml 產生地點報表 Excel。

優點／缺點欄位預留空白，可之後在 Cursor 請 AI 用網頁搜尋整理各店評價，
產出 CSV（店名, 優點, 缺點）後用 merge_pros_cons.py 合併進本 Excel。
"""
import math
import re
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    raise

NS = {"kml": "http://www.opengis.net/kml/2.2"}

# 酒店座標（用於計算距離）；若 KML 有 The Westin 則會覆寫
HOTEL_COORDS = (13.7388, 100.5601)  # lat, lon
# 泰銖 → 港幣（約）；可依匯率調整
THB_TO_HKD = 0.22

def strip_html(html):
    if not html:
        return ""
    text = re.sub(r"<br\s*/?>", "\n", html, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"&nbsp;", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def extract_links(html):
    if not html:
        return []
    urls = re.findall(r'<a\s+href="([^"]+)"[^>]*>([^<]*)</a>', html)
    return urls  # list of (url, text)

def format_links_for_cell(links):
    if not links:
        return ""
    return " | ".join(f"{text} {url}" for url, text in links)

def folder_to_type(folder_name, place_name):
    """依資料夾與名稱推斷分店類型"""
    if "1_酒店與餐廳" in folder_name:
        if "Westin" in place_name or "酒店" in place_name or "Hotel" in place_name:
            return "酒店"
        return "餐廳"
    if "2_日按與Soapy" in folder_name:
        if "日按集中區" in place_name:
            return "日按集中區"
        if "Eden" in place_name:
            return "Soapy／俱樂部"
        if "Emmanuelle" in place_name or "Poseidon" in place_name or "The Lord" in place_name:
            return "泰浴（Soapy）"
        return "日按／Soapy"
    if "3_景點" in folder_name:
        return "景點"
    if "4_紅燈區" in folder_name:
        if "Soi Cowboy" in place_name or "Nana" in place_name:
            return "紅燈區（Go Go Bar）"
        if "Patpong" in place_name:
            return "紅燈區＋夜市"
        return "紅燈區"
    if "5_他人推薦" in folder_name:
        # 從 description 的 分類 擷取
        return "見描述"
    return folder_name.replace("_", " ")

def get_google_maps_url(lat, lon):
    """產生可直接在 Google 地圖開啟的網址（座標查詢）"""
    if lat and lon:
        return f"https://www.google.com/maps?q={lat},{lon}"
    return ""

def get_address_from_desc(desc, coords):
    """地址欄：回傳 Google 地圖可直接使用的網址（點擊或複製貼上皆可開啟）"""
    lat, lon = coords if len(coords) >= 2 else ("", "")
    return get_google_maps_url(lat, lon)

def get_type_from_desc(desc):
    """從他人推薦的 description 擷取 分類"""
    m = re.search(r"<b>分類</b>[：:]\s*([^｜|<]+)", desc)
    if m:
        return m.group(1).strip()
    return ""


def parse_specialty(desc):
    """特色：從 <b>特色</b>、<b>口碑</b> 及描述文字擷取（國籍／年齡／主題／服務等）"""
    parts = []
    for tag in ["特色", "口碑"]:
        m = re.search(rf"<b>{tag}</b>[：:]\s*([^<]+?)(?:<|$)", desc)
        if m:
            parts.append(strip_html(m.group(1)).strip())
    if parts:
        return "；".join(p for p in parts if p)
    # 無標籤時取第一段（介紹連結之前）
    desc_clean = re.sub(r"<b>介紹連結：</b>.*", "", desc, flags=re.DOTALL)
    first = strip_html(desc_clean).strip()
    return first[:200] if first else ""


def _parse_thb_numbers(desc):
    """從描述擷取泰銖數字（含 銖、, 與範圍），回傳 (min_thb, max_thb) 或 None"""
    # 例如 4,500–6,000 銖、約 3,000 銖、30,000／50,000／100,000 銖
    numbers = []
    for m in re.finditer(r"[\d,]+(?:\s*[–\-～~／/]\s*[\d,]+)*\s*銖", desc):
        segment = m.group(0)
        for n in re.findall(r"[\d,]+", segment):
            numbers.append(int(n.replace(",", "")))
    if not numbers:
        return None
    return (min(numbers), max(numbers))


def parse_price_hkd(desc):
    """價錢範圍：從描述擷取泰銖並換算成港幣，回傳如「約 660–1,320 HKD」"""
    thb = _parse_thb_numbers(desc)
    if not thb:
        return ""
    low = int(thb[0] * THB_TO_HKD)
    high = int(thb[1] * THB_TO_HKD)
    if low == high:
        return f"約 {low} HKD"
    return f"約 {low}–{high} HKD"


def parse_hours(desc):
    """開店時間：擷取營業時間與休息日"""
    # 例如 11:00–02:00、12:30-24:00、08:00–24:00、周一休息
    hours = []
    for m in re.finditer(r"\d{1,2}\s*:\s*\d{2}\s*[–\-～~]\s*\d{1,2}\s*:\s*\d{2}", desc):
        hours.append(m.group(0).replace(" ", ""))
    for m in re.finditer(r"[周星期一二三四五六日]+\s*休息", desc):
        hours.append(m.group(0))
    return "；".join(hours) if hours else ""


def parse_rating(desc):
    """分數：擷取 Google 評分，如 4.6★ (77 則)"""
    m = re.search(r"<b>評分</b>[：:]\s*([^<]+)", desc)
    if m:
        return strip_html(m.group(1)).strip()
    m = re.search(r"(\d\.\d)\s*★\s*[（(]?\s*[\d,]+\s*則", desc)
    if m:
        return m.group(0).strip()
    return ""


def haversine_km(lat1, lon1, lat2, lon2):
    """兩點經緯度計算直線距離（公里）"""
    try:
        la1, lo1, la2, lo2 = float(lat1), float(lon1), float(lat2), float(lon2)
    except (TypeError, ValueError):
        return None
    R = 6371  # 地球半徑 km
    phi1, phi2 = math.radians(la1), math.radians(la2)
    dphi = math.radians(la2 - la1)
    dlam = math.radians(lo2 - lo1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return round(R * c, 1)


def parse_remarks(desc):
    """備註：永久歇業、會員制、周一休息、需預約等"""
    remarks = []
    if "永久歇業" in desc or "歇業" in desc:
        remarks.append("永久歇業")
    if "會員制" in desc or "邀請制" in desc:
        remarks.append("會員制")
    if "周一休息" in desc or "星期一休息" in desc:
        remarks.append("周一休息")
    if "預約" in desc and "LINE" in desc:
        remarks.append("可 LINE 預約")
    return "；".join(remarks) if remarks else ""


def parse_transport(desc):
    """最近交通：BTS／MRT／Soi 等"""
    # 從描述擷取 Asok、Nana BTS、Sutthisan MRT、Phrom Phong、Soi XX 等
    tokens = []
    for m in re.finditer(r"(?:Asok|Nana|Phrom Phong|Ekkamai|Thong Lo)\s*(?:BTS)?|Sutthisan\s*MRT|Soi\s*\d+(?:/\d+)?|Ratchada|RCA", desc, re.I):
        tokens.append(m.group(0).strip())
    return "、".join(dict.fromkeys(tokens)) if tokens else ""


def main():
    kml_path = Path(__file__).parent / "地圖.kml"
    tree = ET.parse(kml_path)
    root = tree.getroot()

    hotel_coords = HOTEL_COORDS  # (lat, lon)
    rows = []
    for folder in root.findall(".//kml:Folder", NS):
        folder_name_el = folder.find("kml:name", NS)
        folder_name = folder_name_el.text if folder_name_el is not None else ""

        for pm in folder.findall(".//kml:Placemark", NS):
            name_el = pm.find("kml:name", NS)
            name = name_el.text or ""

            desc_el = pm.find("kml:description", NS)
            desc = (desc_el.text or "") if desc_el is not None else ""

            coords_el = pm.find(".//kml:coordinates", NS)
            if coords_el is not None and coords_el.text:
                parts = coords_el.text.strip().split(",")
                coords = (parts[1], parts[0]) if len(parts) >= 2 else ("", "")
            else:
                coords = ("", "")

            # 辨識酒店並記錄座標（用於距離計算）
            if "1_酒店與餐廳" in folder_name and "Westin" in name:
                hotel_coords = (float(coords[0]), float(coords[1])) if coords[0] and coords[1] else hotel_coords

            if "5_他人推薦" in folder_name:
                c_type = get_type_from_desc(desc) or "成人按摩／俱樂部"
            else:
                c_type = folder_to_type(folder_name, name)

            address = get_address_from_desc(desc, coords)
            links = extract_links(desc)
            d_links = format_links_for_cell(links)

            specialty = parse_specialty(desc)
            price_hkd = parse_price_hkd(desc)
            hours = parse_hours(desc)
            rating = parse_rating(desc)
            remarks = parse_remarks(desc)
            transport = parse_transport(desc)

            # 距離：與酒店直線距離（公里）；酒店本身顯示「—」
            if "Westin" in name and "1_酒店" in folder_name:
                distance = "—"
            elif coords[0] and coords[1]:
                km = haversine_km(hotel_coords[0], hotel_coords[1], coords[0], coords[1])
                distance = f"約 {km} km" if km is not None else ""
            else:
                distance = ""

            rows.append({
                "A名稱": name,
                "B地址": address,
                "C分店類型": c_type,
                "D相關介紹網頁": d_links,
                "特色": specialty,
                "價錢範圍(港幣)": price_hkd,
                "開店時間": hours,
                "分數": rating,
                "優點": "",
                "缺點": "",
                "距離": distance,
                "備註": remarks,
                "最近交通": transport,
            })

    # 寫入 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "地點報表"
    headers = [
        "A名稱", "B地址", "C分店類型", "D相關介紹網頁",
        "特色", "價錢範圍(港幣)", "開店時間", "分數", "優點", "缺點", "距離",
        "備註", "最近交通",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r["A名稱"])
        addr_cell = ws.cell(row=row_idx, column=2, value=r["B地址"])
        if r["B地址"]:
            addr_cell.hyperlink = r["B地址"]
            addr_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=row_idx, column=3, value=r["C分店類型"])
        ws.cell(row=row_idx, column=4, value=r["D相關介紹網頁"])
        ws.cell(row=row_idx, column=5, value=r["特色"])
        ws.cell(row=row_idx, column=6, value=r["價錢範圍(港幣)"])
        ws.cell(row=row_idx, column=7, value=r["開店時間"])
        ws.cell(row=row_idx, column=8, value=r["分數"])
        ws.cell(row=row_idx, column=9, value=r["優點"])
        ws.cell(row=row_idx, column=10, value=r["缺點"])
        ws.cell(row=row_idx, column=11, value=r["距離"])
        ws.cell(row=row_idx, column=12, value=r["備註"])
        ws.cell(row=row_idx, column=13, value=r["最近交通"])
        for c in range(1, 14):
            ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 28
    ws.column_dimensions["J"].width = 28
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 16
    ws.column_dimensions["M"].width = 22

    out_path = Path(__file__).parent / "地點報表.xlsx"
    wb.save(out_path)
    print(f"已產生: {out_path}")

if __name__ == "__main__":
    main()
